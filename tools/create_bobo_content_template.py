from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "content"
OUT_FILE = OUT_DIR / "bobo_content_quick_template.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="163A5F")
EXAMPLE_FILL = PatternFill("solid", fgColor="EAF7F4")
NOTE_FILL = PatternFill("solid", fgColor="FFF6D7")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="163A5F")
WRAP = Alignment(wrap_text=True, vertical="top")


def add_validation(ws, col_letter, values, start=2, end=5000):
    quoted = ",".join(values)
    dv = DataValidation(type="list", formula1=f'"{quoted}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start}:{col_letter}{end}")


def setup_sheet(ws, headers, descriptions, example):
    ws.freeze_panes = "A2"
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.append(example)
    for cell in ws[2]:
        cell.fill = EXAMPLE_FILL
        cell.alignment = WRAP
    for idx, header in enumerate(headers, 1):
        letter = ws.cell(row=1, column=idx).column_letter
        ws.cell(row=1, column=idx).comment = Comment(descriptions.get(header, ""), "Codex")
        max_len = max(len(str(header)), len(str(example[idx - 1])) if idx - 1 < len(example) else 10)
        ws.column_dimensions[letter].width = min(max(max_len + 4, 14), 44)
    ws.auto_filter.ref = ws.dimensions


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    readme = wb.create_sheet("README")
    readme["A1"] = "BoBo English 内容后台模板"
    readme["A1"].font = TITLE_FONT
    rows = [
        ("使用顺序", "先填 units，再填 words，然后填 phrases / passages / word_tests。"),
        ("最重要", "words 里的 en、zh、app_sentence、example_1、example_2 请尽量使用原创内容，避免版权风险。"),
        ("category", "建议只用 5 类：核心詞彙、高頻基礎詞、考試重點詞、短語搭配、拓展詞彙。"),
        ("source_note", "可选。用于内部追踪来源，例如 Teacher-created、Text 2、Unit Assessment。不填也可以。"),
        ("word_tests", "用于第二轮/第三轮考点题，一行就是一道题。weakness_tag 用来统计孩子薄弱点。"),
        ("enabled", "TRUE 表示启用，FALSE 表示暂时不用。"),
    ]
    for r, (k, v) in enumerate(rows, 3):
        readme.cell(r, 1, k).font = Font(bold=True, color="163A5F")
        readme.cell(r, 2, v).alignment = WRAP
    readme.column_dimensions["A"].width = 18
    readme.column_dimensions["B"].width = 92

    quick_headers = [
        "unit_key", "form", "book", "unit_num", "theme_en", "theme_zh",
        "group_no", "word_order", "en", "pos", "zh", "source_sentence",
        "focus", "notes", "enabled"
    ]
    quick_desc = {
        "unit_key": "Required. Unit id, e.g. 1A_U1. Same unit should use the same value.",
        "form": "Required. F1 / F2 / F3.",
        "book": "Required. Book label, e.g. 1A, 1B, 2A.",
        "unit_num": "Required. Unit number only, e.g. 1.",
        "theme_en": "Required. Unit theme in English.",
        "theme_zh": "Required. Unit theme in Chinese.",
        "group_no": "Required. Word group number. One group can contain 9, 15, or 20 words.",
        "word_order": "Required. Word order inside the group.",
        "en": "Required. English word or phrase.",
        "pos": "Required. Part of speech: n. / v. / adj. / adv. / phr. etc.",
        "zh": "Required. Chinese meaning.",
        "source_sentence": "Optional. Original/source sentence or your rough sentence. Codex can rewrite it into copyright-safe examples.",
        "focus": "Optional. What you want to test: word_form / plural / tense / preposition / collocation / spelling etc.",
        "notes": "Optional. Any teacher notes, source notes, or special requirements.",
        "enabled": "Required. TRUE to use, FALSE to ignore.",
    }
    quick_example = [
        "1A_U1", "F1", "1A", 1, "Me and my world", "我與我的世界",
        1, 1, "campus", "n.", "校園",
        "On campus, there is a huge playground for all kinds of ball games.",
        "preposition; collocation", "Rewrite source sentence; avoid textbook wording.", "TRUE"
    ]
    ws = wb.create_sheet("quick_words")
    setup_sheet(ws, quick_headers, quick_desc, quick_example)
    add_validation(ws, "B", ["F1", "F2", "F3"])
    add_validation(ws, "J", ["n.", "v.", "adj.", "adv.", "phr.", "prep.", "conj."])
    add_validation(ws, "O", ["TRUE", "FALSE"])

    units_headers = ["unit_key", "form", "book", "unit_num", "theme_en", "theme_zh", "display_order", "enabled"]
    units_desc = {
        "unit_key": "单元唯一编号，不要重复。建议格式：1A_U1、1B_U2。",
        "form": "年级阶段：F1、F2、F3。",
        "book": "教材册别，例如 1A、1B、2A。",
        "unit_num": "第几单元，只填数字。",
        "theme_en": "单元英文主题。",
        "theme_zh": "单元中文主题。",
        "display_order": "App 中显示顺序，数字越小越靠前。",
        "enabled": "TRUE 启用，FALSE 暂停。",
    }
    units_example = ["1A_U1", "F1", "1A", 1, "Me and my world", "我與我的世界", 1, "TRUE"]
    ws = wb.create_sheet("units")
    setup_sheet(ws, units_headers, units_desc, units_example)
    add_validation(ws, "B", ["F1", "F2", "F3"])
    add_validation(ws, "H", ["TRUE", "FALSE"])

    words_headers = [
        "unit_key", "group_no", "word_order", "en", "pos", "zh", "category", "source_note",
        "app_sentence", "exam_tip", "error_tip", "collocations", "example_1", "example_2", "enabled"
    ]
    words_desc = {
        "unit_key": "所属单元，必须和 units 表一致。",
        "group_no": "第几组。建议每组 9 个词。",
        "word_order": "组内顺序，建议 1-9。",
        "en": "英文单词或词组。",
        "pos": "词性，例如 n. / v. / adj. / adv. / phr.",
        "zh": "中文意思。",
        "category": "词汇类别，建议从下拉 5 类中选。",
        "source_note": "可选。内部追踪来源，例如 Teacher-created、Text 2、Unit Assessment。",
        "app_sentence": "App 出题用原创例句。",
        "exam_tip": "可选。考试/题型提示。",
        "error_tip": "易错提醒。学生答错后可显示。",
        "collocations": "常见搭配，用分号隔开。",
        "example_1": "扩展原创例句 1。",
        "example_2": "扩展原创例句 2。",
        "enabled": "TRUE 启用，FALSE 暂停。",
    }
    words_example = [
        "1A_U1", 3, 6, "ambitious", "adj.", "有抱負的；野心勃勃的", "考試重點詞",
        "Teacher-created", "Mandy is ambitious and hopes to lead the school team one day.",
        "常考性格形容詞；也可表示計劃過於進取。",
        "注意 ambitious 多為褒義；拼寫是 -tious，不是 -cious。",
        "an ambitious student; ambitious plans; be ambitious for the future",
        "He is ambitious, so he practises speaking English every day.",
        "It is too ambitious to finish five projects in one weekend.",
        "TRUE",
    ]
    ws = wb.create_sheet("words")
    setup_sheet(ws, words_headers, words_desc, words_example)
    add_validation(ws, "E", ["n.", "v.", "adj.", "adv.", "phr.", "prep.", "conj."])
    add_validation(ws, "G", ["核心詞彙", "高頻基礎詞", "考試重點詞", "短語搭配", "拓展詞彙"])
    add_validation(ws, "O", ["TRUE", "FALSE"])

    phrases_headers = ["unit_key", "phrase_order", "phrase", "zh", "type", "app_sentence", "note", "example", "enabled"]
    phrases_desc = {
        "unit_key": "所属单元，必须和 units 表一致。",
        "phrase_order": "短语显示顺序。",
        "phrase": "英文短语。",
        "zh": "中文意思。",
        "type": "短语类型。",
        "app_sentence": "App 出题用原创例句。",
        "note": "用法说明或易错点。",
        "example": "扩展原创例句。",
        "enabled": "TRUE 启用，FALSE 暂停。",
    }
    phrases_example = [
        "1A_U1", 1, "lend a hand", "幫忙", "短語搭配",
        "The librarian lent me a hand when I could not find the book.",
        "lend sb a hand = help sb",
        "My classmate lent me a hand with the group project.",
        "TRUE",
    ]
    ws = wb.create_sheet("phrases")
    setup_sheet(ws, phrases_headers, phrases_desc, phrases_example)
    add_validation(ws, "E", ["短語搭配", "口語常用", "寫作句式", "固定句型", "補充短語"])
    add_validation(ws, "I", ["TRUE", "FALSE"])

    passages_headers = ["unit_key", "group_no", "title", "target_words", "passage_text", "question_mode", "enabled"]
    passages_desc = {
        "unit_key": "所属单元，必须和 units 表一致。",
        "group_no": "对应第几组词。",
        "title": "短文标题。",
        "target_words": "短文重点词，用英文逗号分隔。",
        "passage_text": "原创短文。建议 80-150 词，逻辑清楚，有故事性。",
        "question_mode": "短文题型。",
        "enabled": "TRUE 启用，FALSE 暂停。",
    }
    passages_example = [
        "1A_U1", 3, "A New Goal",
        "ambitious,determined,outstanding,cheerful",
        "Mandy used to be quiet in class, but she had an ambitious goal. She wanted to speak English clearly in front of the whole school. Every day, she stayed determined and practised with her cheerful classmates. By the end of the term, her performance was outstanding.",
        "fill_blank",
        "TRUE",
    ]
    ws = wb.create_sheet("passages")
    setup_sheet(ws, passages_headers, passages_desc, passages_example)
    add_validation(ws, "F", ["fill_blank", "choose_word", "word_form", "mixed"])
    add_validation(ws, "G", ["TRUE", "FALSE"])

    tests_headers = [
        "unit_key", "group_no", "en", "test_type", "question",
        "option_a", "option_b", "option_c", "option_d", "answer",
        "explanation", "weakness_tag", "difficulty", "enabled"
    ]
    tests_desc = {
        "unit_key": "所属单元，必须和 units 表一致。",
        "group_no": "对应第几组词。",
        "en": "这道题对应的目标单词。",
        "test_type": "考点类型，从下拉选择。",
        "question": "题目句子或题干。",
        "option_a": "选项 A。",
        "option_b": "选项 B。",
        "option_c": "选项 C。",
        "option_d": "选项 D。",
        "answer": "正确答案。必须和某个选项完全一致。",
        "explanation": "答错后的解释。",
        "weakness_tag": "错了说明哪方面薄弱，用于学生报告。",
        "difficulty": "难度：easy / medium / hard。",
        "enabled": "TRUE 启用，FALSE 暂停。",
    }
    tests_example = [
        "1A_U1", 3, "ambitious", "word_form",
        "She has a strong ____ to become a doctor.",
        "ambitious", "ambition", "ambitiously", "ambitions", "ambition",
        "这里需要名词，ambition 表示“抱负”。",
        "詞性變化", "medium", "TRUE",
    ]
    ws = wb.create_sheet("word_tests")
    setup_sheet(ws, tests_headers, tests_desc, tests_example)
    add_validation(ws, "D", [
        "word_form", "plural", "tense", "preposition", "collocation", "countability",
        "ed_ing", "gerund_infinitive", "subject_verb_agreement", "article",
        "confusing_words", "context_meaning", "spelling", "writing_upgrade"
    ])
    add_validation(ws, "L", [
        "詞性變化", "名詞單複數", "動詞時態", "介詞搭配", "固定搭配",
        "可數不可數", "ed/ing形容詞", "動名詞/不定式", "主謂一致",
        "冠詞", "易混詞", "語境判斷", "拼寫", "寫作替換"
    ])
    add_validation(ws, "M", ["easy", "medium", "hard"])
    add_validation(ws, "N", ["TRUE", "FALSE"])

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = WRAP
        if ws.max_row >= 2:
            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 72

    readme["A10"] = "提示"
    readme["A10"].font = TITLE_FONT
    readme["B10"] = "第 2 行是示例，可以复制参考；正式导入时我会忽略或删除示例行。"
    readme["B10"].fill = NOTE_FILL
    readme["B10"].alignment = WRAP

    wb.save(OUT_FILE)
    print(OUT_FILE)


if __name__ == "__main__":
    main()
